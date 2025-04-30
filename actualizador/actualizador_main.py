import os
import io
import logging  # Importar el módulo logging
import pandas as pd
from openpyxl.writer.excel import save_virtual_workbook
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from bdd.models import Listado_Planillas, Item, Sub_Carpeta, Sub_Titulo, ListaProveedores
from bdd.classes import Patoba
from bdd.funtions import get_emails
from django.conf import settings as const # Corregido: cosnt -> const
from asgiref.sync import sync_to_async
# from django.conf import settings # Ya importado como const
import json
import openpyxl
import csv
from openpyxl.utils.dataframe import dataframe_to_rows # Import necesario para f() si se usa
from x_cartel.models import Carteles, CartelesCajon


# Configuración básica de logging (puedes personalizarla más)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# --- Función registrar_log sin modificar ---
# Esta función escribe a un archivo 'log.txt' específico.
# No se reemplaza con logger para mantener la lógica original de escribir a ese archivo.
def registrar_log(texto):
    try:
        with open('log.txt', 'a', encoding='utf-8') as f: # Añadir encoding
            print(texto, file=f)
    except Exception as e:
        # Loguear error si no se puede escribir en log.txt
        logger.error(f"No se pudo escribir en log.txt: {e}")

def crear_o_actualizar_registro(row):
    item = None # Inicializar item
    codigo_item = row.get('codigo', 'SIN CODIGO') # Obtener código para logging
    try:
        # Recuperar o crear una instancia de Sub_Carpeta
        sub_carpeta_nombre = row.pop('sub_carpeta', None) # Usar pop con default None
        if sub_carpeta_nombre:
            sub_carpeta, created = Sub_Carpeta.objects.get_or_create(nombre=sub_carpeta_nombre)
            row['sub_carpeta'] = sub_carpeta
        else:
            row['sub_carpeta'] = None # Asignar None si no viene

        # Recuperar o crear una instancia de Sub_Titulo
        sub_titulo_nombre = row.pop('sub_titulo', None) # Usar pop con default None
        if sub_titulo_nombre:
            sub_titulo, created = Sub_Titulo.objects.get_or_create(nombre=sub_titulo_nombre)
            row['sub_titulo'] = sub_titulo
        else:
            row['sub_titulo'] = None # Asignar None si no viene

    except Exception as e:
        # Reemplazo de print(e)
        logger.error(f"Error al procesar Sub_Carpeta/Sub_Titulo para código {codigo_item}: {e}", exc_info=True)
        # Considerar si retornar aquí o continuar sin carpeta/título

    # logger.debug(f"Fila procesada antes de guardar para código {codigo_item}: {row}") # Reemplazo de #print(row)

    try:
        # Usar update_or_create es generalmente más eficiente y atómico
        # Pero manteniendo la lógica original con get_or_create y luego setattr/save:
        item, created = Item.objects.get_or_create(codigo=codigo_item, defaults=row)

        if not created:
            # Actualizar campos existentes si no fue creado
            for key, value in row.items():
                # logger.debug(f"Actualizando item {codigo_item}: {key}={value}") # Reemplazo de #print(item, key, value)
                # Solo actualizar si el valor es diferente para evitar saves innecesarios
                if hasattr(item, key) and getattr(item, key) != value:
                   setattr(item, key, value)
            # Guardar una sola vez después del bucle si no fue creado y hubo cambios (o siempre si la lógica original lo hacía)
            item.save() # Moviendo save fuera del bucle
        # else: # Si fue creado, los defaults ya se aplicaron, solo falta guardar y marcar
        #     item.save() # Si defaults no guarda automáticamente

    except Exception as e:
         # Reemplazo de print(e)
        logger.error(f"Error durante get_or_create o actualización de atributos para código {codigo_item}: {e}", exc_info=True)
        # No continuar si falla la creación/obtención

    # Guardar y marcar actualizado después de crear o actualizar atributos
    try:
        if item: # Asegurarse que item existe
            item.marcar_actualizado() # Marcar actualizado
            item.save() # Guardar el estado actualizado y cualquier cambio de setattr
            # logger.info(f"{'Creado' if created else 'Actualizado'} y guardado: {codigo_item}") # Reemplazo de #print('creado o guardado: ', item)
        else:
             logger.warning(f"No se pudo guardar o marcar actualizado el item {codigo_item} porque no se creó/obtuvo correctamente.")

    except Exception as e:
         # Reemplazo de print("item no se guardo...") y print(e)
        logger.error(f"Item no se guardó o marcó actualizado debido a error final. Código: {codigo_item}. Fila: {row}", exc_info=True)


def desactualizar_anteriores(filtro):
    """Marca como no actualizados los items cuyo código termina con el filtro dado."""
    logger.info(f"Marcando como no actualizados items terminados en: {filtro}")
    try:
        # La siguiente línea YA USA update(), que es eficiente.
        if filtro: # Asegurar que el filtro no sea vacío o None
            # en la base de datos para todos los objetos que coinciden con el filtro,
            count = Item.objects.filter(codigo__endswith=filtro).update(actualizado=False)
            logger.info(f"{count} items marcados como no actualizados para el filtro: {filtro}")
        else:
            logger.warning("Se intentó desactualizar con un filtro vacío o None.")
    except Exception as e:
        logger.error(f"Error al desactualizar items con filtro '{filtro}': {e}", exc_info=True)

def buscar_modificar_registros(csv_file, filtro):
    """Procesa un archivo CSV para crear o actualizar registros uno por uno."""
    logger.info(f"Iniciando procesamiento individual para archivo: {csv_file}, filtro: {filtro}")
    try:
        desactualizar_anteriores(filtro)
        # Reemplazo de print('desactualizado: ',filtro) - logging ya está en desactualizar_anteriores

        processed_count = 0
        error_count = 0
        skipped_count = 0

        with open(csv_file, newline='', encoding='utf-8') as csvfile: # Añadir encoding
            reader = csv.DictReader(csvfile)
            for row in reader:
                # Limpiar espacios en claves y valores
                row = {k.strip() if k else k: v.strip() if isinstance(v, str) else v for k, v in row.items() if k} # Ignorar claves vacías
                codigo_item = row.get('codigo', 'SIN CODIGO')
                # logger.debug(f"Procesando fila (individual) código {codigo_item}: {row}") # Reemplazo de print(row)

                precio_base = row.get('precio_base')
                invalid_price_values = [None, '', ' ', '-', '.', '#N/A', '#VALUE!']

                # Filtrar filas con precio_base inválido
                if precio_base in invalid_price_values:
                    # logger.debug(f"Saltando fila por precio_base inválido ('{precio_base}') para código {codigo_item}")
                    skipped_count += 1
                    continue

                try:
                    crear_o_actualizar_registro(row)
                    processed_count += 1
                except Exception as e:
                    # La función crear_o_actualizar_registro ya loguea sus errores internos.
                    # Podemos añadir un log aquí si queremos saber que falló en este nivel.
                    logger.error(f"Error llamando a crear_o_actualizar_registro para código {codigo_item} desde buscar_modificar_registros.", exc_info=True)
                    error_count += 1

        logger.info(f"Procesamiento individual completado para {filtro}. Procesados: {processed_count}, Errores: {error_count}, Saltados: {skipped_count}")
        # Reemplazo de print('cargado: ', filtro)

    except FileNotFoundError:
        logger.error(f"Archivo CSV no encontrado: {csv_file}")
    except Exception as e:
        logger.error(f"Error general en buscar_modificar_registros para {csv_file} y filtro {filtro}: {e}", exc_info=True)

# --- Función de fusión de hojas ---
def fusionar_hojas(file_stream):
    """Carga un stream de archivo Excel, fusiona hojas (excepto 'general') en 'cargar datos' y devuelve el Workbook."""
    logger.debug("Iniciando fusión de hojas...")
    try:
        # Cargar el archivo xlsx en un objeto Workbook de openpyxl
        wb = openpyxl.load_workbook(file_stream)

        # Filtrar hojas a fusionar
        hojas_a_fusionar = [ws for ws in wb.worksheets if ws.title != "general"]
        if not hojas_a_fusionar:
            logger.warning("No se encontraron hojas para fusionar (excluyendo 'general').")
            # Decide qué hacer: devolver wb tal cual, o un wb con 'cargar datos' vacía?
            # Por ahora, crearemos 'cargar datos' si no existe y la dejaremos vacía.
            if "cargar datos" not in wb.sheetnames:
                wb.create_sheet(title="cargar datos")
            return wb

        # Fusionar todas las hojas filtradas en una sola DataFrame
        df = pd.concat([pd.DataFrame(ws.values) for ws in hojas_a_fusionar])
        logger.debug(f"DataFrame fusionado creado con {len(df)} filas.")

        # Crear o seleccionar la hoja "cargar datos"
        if "cargar datos" in wb.sheetnames:
            ws = wb["cargar datos"]
            # Opcional: Limpiar hoja existente antes de escribir
            # ws.delete_rows(1, ws.max_row)
            logger.debug("Usando hoja 'cargar datos' existente.")
        else:
            ws = wb.create_sheet(title="cargar datos")
            logger.debug("Creando nueva hoja 'cargar datos'.")


        # Convertir el DataFrame y agregar a la hoja
        # Usar dataframe_to_rows para mejor manejo de tipos
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1): # header=True para incluir encabezado si df lo tiene
             # Escribir fila - ajustar según si quieres header o no
             # El append directo puede ser más simple si no necesitas control fino
             ws.append(row)
             # logger.debug(f"Fila {r_idx} añadida a 'cargar datos'.") # Muy verboso

        logger.info("Hojas fusionadas exitosamente en la hoja 'cargar datos'.")
        # Retornar el objeto Workbook modificado
        return wb

    except Exception as e:
        logger.error(f"Error durante la fusión de hojas: {e}", exc_info=True)
        raise # Relanzar la excepción para que el llamador sepa que falló

def marcar_revisar_carteles(id_proveedor):
    """Marca los carteles y carteles de cajón asociados a un proveedor para revisión."""
    try:
        # logger.info(f"Marcando carteles para revisar, proveedor ID: {id_proveedor}") # Reemplazo de print
        count1 = Carteles.objects.filter(proveedor=id_proveedor).update(revisar=True)
        count2 = CartelesCajon.objects.filter(proveedor=id_proveedor).update(revisar=True)
        logger.info(f"Marcados para revisión: {count1} carteles y {count2} carteles de cajón para proveedor ID: {id_proveedor}")
    except Exception as e:
        logger.error(f"Error al marcar carteles para revisar para proveedor ID {id_proveedor}: {e}", exc_info=True)


def principal():
    """Función principal del script para procesar planillas de Google Drive/Sheets."""
    logger.info("Inicio del proceso principal.")
    patoba = Patoba(None) # Asume inicialización correcta

    # Verificar servicios
    if not patoba.gmail_service or not patoba.drive_service or not patoba.sheet_service:
        logger.critical("No se pudieron inicializar los servicios de Google. Abortando.")
        return

    gmail_service = patoba.gmail_service
    drive_service = patoba.drive_service
    sheet_service = patoba.sheet_service # Añadido para claridad

    # --- Procesamiento de Emails (si aplica) ---
    try:
        logger.info("Buscando y procesando emails...")
        get_emails(gmail_service, drive_service) # Asume que esta función existe y loguea internamente
        logger.info("Procesamiento de emails completado.")
    except Exception as e:
        logger.error(f"Error durante get_emails: {e}", exc_info=True)

    # --- Listar y Crear/Actualizar Entradas de Planillas en BD ---
    try:
        logger.info(f"Listando archivos en la carpeta INBOX ({const.INBOX})...")
        # Asegúrate que const.INBOX esté definido en tus settings
        items_drive = patoba.listar(100, const.INBOX) # Listar archivos de Drive
        logger.info(f"Encontrados {len(items_drive)} archivos en INBOX.")
        created_count = 0
        updated_count = 0
        error_count = 0
        for item_drive in items_drive:
            try:
                # Usar update_or_create para manejar creación y actualización
                obj, created = Listado_Planillas.objects.update_or_create(
                    identificador=item_drive['id'],
                    defaults={'descripcion': item_drive['name']}
                )
                if created:
                    created_count += 1
                else:
                    updated_count +=1
            except Exception as e:
                 # Reemplazo de print('No se pudo get_or_create...')
                logger.error(f"No se pudo update_or_create Listado_Planillas para [{item_drive.get('name', 'N/A')}] , [{item_drive.get('id', 'N/A')}]", exc_info=True)
                error_count += 1
        logger.info(f"Planillas en BD: {created_count} creadas, {updated_count} actualizadas/encontradas, {error_count} errores.")

    except Exception as e:
        logger.error(f"Error listando archivos de Drive o actualizando Listado_Planillas: {e}", exc_info=True)


    # --- Procesamiento específico para 'Caramia' (Fusión de Hojas) ---
    try:
        logger.info("Procesando caso especial: Caramia (fusión de hojas)...")
        # Usar filter().first() para manejar el caso de que no exista
        planilla_caramia = Listado_Planillas.objects.filter(proveedor__text_display="Caramia").first()

        if planilla_caramia:
            file_id_caramia = planilla_caramia.identificador
            logger.info(f"Descargando archivo de Caramia (ID: {file_id_caramia})...")
            request = drive_service.files().get_media(fileId=file_id_caramia)
            file_stream_caramia = io.BytesIO()
            downloader = MediaIoBaseDownload(file_stream_caramia, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()
                # logger.debug(f"Descarga Caramia: Progreso {int(status.progress() * 100)}%")

            file_stream_caramia.seek(0)
            logger.info("Archivo Caramia descargado. Fusionando hojas...")
            # Llamar a la función de fusión
            wb_caramia = fusionar_hojas(file_stream_caramia) # Pasa el stream

            # Guardar los cambios en Google Drive
            logger.info(f"Guardando archivo modificado de Caramia (ID: {file_id_caramia}) en Drive...")
            file_content_caramia = io.BytesIO(save_virtual_workbook(wb_caramia))
            media = MediaIoBaseUpload(file_content_caramia, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
            drive_service.files().update(fileId=file_id_caramia, media_body=media).execute()
            logger.info("Archivo Caramia actualizado en Drive con hoja 'cargar datos'.")
            # Considerar si actualizar planilla_caramia.hoja en la BD aquí
            # planilla_caramia.hoja = "cargar datos"
            # planilla_caramia.save()

        else:
            logger.warning("No se encontró registro de planilla para Caramia en la base de datos.")

    except Exception as e:
        # Reemplazo de print("no se encontro caramia {}".format(e), flush=True)
        logger.error(f"Error procesando el caso especial de Caramia: {e}", exc_info=True)


    # --- Obtener Nombres de Hojas para Planillas no listas ---
    try:
        planillas_no_listas = Listado_Planillas.objects.filter(listo=False, descargar=False)
        logger.info(f"Verificando hojas para {planillas_no_listas.count()} planillas marcadas como no listas/no descargar...")
        for planilla in planillas_no_listas:
            logger.debug(f"Obteniendo hojas para: {planilla.descripcion} (ID: {planilla.identificador})")
            try:
                request = drive_service.files().get_media(fileId=planilla.identificador)
                file = io.BytesIO()
                downloader = MediaIoBaseDownload(file, request)
                done = False
                while done is False:
                    status, done = downloader.next_chunk()
                file.seek(0)

                # Leer nombres de hojas
                # Usar openpyxl es más ligero que pandas solo para nombres de hojas
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                sheet_names = wb.sheetnames # Obtiene lista de nombres
                wb.close() # Cerrar el workbook

                # Guardar nombres en BD (ejemplo: separados por ';')
                planilla.hojas = ';'.join(sheet_names)
                # logger.info(f"Hojas encontradas para {planilla.descripcion}: {planilla.hojas}") # Reemplazo de print(dato.hojas, flush=True)
                planilla.save()
                logger.debug(f"Nombres de hojas guardados para {planilla.descripcion}.")

            except Exception as e:
                # Reemplazo de print("error con pandas {}".format(e), flush=True)
                # El error podría ser de descarga o de lectura de hojas
                 logger.error(f"Error obteniendo/guardando hojas para {planilla.descripcion} (ID: {planilla.identificador}): {e}", exc_info=True)
    except Exception as e:
         logger.error(f"Error general en la sección de obtención de nombres de hojas: {e}", exc_info=True)

    # --- Procesamiento Principal de Planillas Marcadas como Listas ---
    try:
        planillas_listas = Listado_Planillas.objects.filter(listo=True)
        logger.info(f"Iniciando procesamiento principal para {planillas_listas.count()} planillas marcadas como listas.")
        mi_diccionario = {} # Para almacenar mapeo nombre_descargable -> identificador

        for planilla_lista in planillas_listas:
            # Renombrar 'sp' a 'planilla_lista' para claridad
            nombre_proveedor_desc = planilla_lista.descripcion # Nombre del archivo descargado
            proveedor_obj = planilla_lista.proveedor # Objeto Proveedor (relacionado via FK)
            nombre_plantilla = proveedor_obj.text_display if proveedor_obj else None # Nombre de la plantilla (desde Proveedor)

            logger.info(f"Procesando: {nombre_proveedor_desc} (Plantilla: {nombre_plantilla})")

            if not nombre_plantilla or nombre_plantilla == 'Otros':
                logger.warning(f"Saltando {nombre_proveedor_desc}: Sin plantilla asociada o es 'Otros'.")
                continue

            nombre_descargable = f"{nombre_plantilla}-{planilla_lista.fecha}" # Usar nombre de plantilla
            hoja_seleccionada = planilla_lista.hoja
            logger.info(f"Hoja seleccionada: {hoja_seleccionada}")

            # --- Obtener IDs ---
            id_archivo_proveedor = planilla_lista.identificador # ID del archivo original descargado
            id_archivo_plantilla = None
            try:
                # Asumiendo que nombre_plantilla es el nombre del archivo de plantilla en Drive
                # y const.PLANTILLAS es el ID de la carpeta de plantillas
                id_archivo_plantilla = patoba.obtener_id_por_nombre(nombre_plantilla, const.PLANTILLAS)
                if not id_archivo_plantilla:
                     logger.error(f"No se encontró ID para la plantilla '{nombre_plantilla}' en la carpeta PLANTILLAS ({const.PLANTILLAS}). Saltando.")
                     continue
                planilla_lista.id_sp = id_archivo_plantilla # Guardar ID de plantilla en el registro
                planilla_lista.save()
                logger.debug(f"ID de archivo proveedor: {id_archivo_proveedor}, ID de plantilla: {id_archivo_plantilla}")
            except Exception as e:
                 logger.error(f"Error obteniendo ID de la plantilla '{nombre_plantilla}': {e}", exc_info=True)
                 continue # Saltar si no podemos obtener el ID

            # --- Exportar BDD de Plantilla a CSV ---
            try:
                logger.info(f"Exportando hoja 'BDD' de la plantilla '{nombre_plantilla}' (ID: {id_archivo_plantilla}) a CSV...")
                result = sheet_service.spreadsheets().values().get(
                    spreadsheetId=id_archivo_plantilla, range='BDD' # Asume que siempre se llama 'BDD'
                ).execute()
                values = result.get('values', [])

                if not values:
                     logger.warning(f"La hoja 'BDD' en la plantilla '{nombre_plantilla}' está vacía o no existe.")
                     # Decide si continuar sin CSV o parar
                else:
                    # Obtener abreviatura del proveedor para el filtro
                    abreviatura_filtro = proveedor_obj.identificador.abreviatura if hasattr(proveedor_obj, 'identificador') and proveedor_obj.identificador else None
                    if not abreviatura_filtro:
                         logger.error(f"No se pudo obtener la abreviatura para el proveedor '{nombre_plantilla}'. No se puede procesar el CSV.")
                         # Decide si continuar o saltar
                    else:
                        logger.info(f"Abreviatura para filtro CSV: {abreviatura_filtro}")
                        csv_filename = f'{nombre_plantilla}.csv'
                        # Guardar los datos en un archivo CSV localmente
                        logger.debug(f"Guardando datos en {csv_filename}...")
                        with open(csv_filename, 'w', newline='', encoding='utf-8') as f:
                            # logger.debug("Punto de control CSV 1") # Reemplazo print
                            writer = csv.writer(f)
                            # logger.debug("Punto de control CSV 2") # Reemplazo print
                            writer.writerows(values)
                        # logger.debug("Punto de control CSV 3") # Reemplazo print
                        logger.info(f"Archivo CSV '{csv_filename}' guardado con {len(values)} filas.")

                        # --- Procesar el CSV (Desactualizar y Cargar) ---
                        logger.info(f"Procesando archivo CSV: {csv_filename}, Filtro: {abreviatura_filtro}")
                        # Llamar a la función que desactualiza y carga (usa la función ya existente)
                        buscar_modificar_registros(csv_filename, abreviatura_filtro)
                        # Marcar carteles para revisión después de procesar el CSV
                        marcar_revisar_carteles(proveedor_obj.id) # Pasa el ID del objeto Proveedor

            except Exception as e:
                # Reemplazo de print("No se encontro BDD", flush=True) y print(e)
                logger.error(f"Error procesando BDD o CSV para plantilla '{nombre_plantilla}' (ID: {id_archivo_plantilla}): {e}", exc_info=True)
                # Considerar si continuar con las siguientes etapas si el CSV falla

            # --- Copiar Hoja 'Reemplazable' ---
            try:
                logger.info(f"Copiando hoja '{hoja_seleccionada}' del archivo descargado a 'Reemplazable' en la plantilla...")
                id_hoja_reemplazable = patoba.obtener_id_hoja_por_nombre("Reemplazable", id_archivo_plantilla)
                if id_hoja_reemplazable is None:
                    logger.error(f"No se encontró la hoja 'Reemplazable' en la plantilla '{nombre_plantilla}' (ID: {id_archivo_plantilla}). No se puede copiar.")
                else:
                    # La función copiar_reemplazable necesita los IDs y el nombre de la hoja origen
                    patoba.copiar_reemplazable(
                        id_archivo_proveedor, # ID del archivo fuente (el descargado)
                        hoja_seleccionada,   # Nombre de la hoja fuente
                        id_hoja_reemplazable,# ID de la hoja destino ('Reemplazable')
                        id_archivo_plantilla # ID del archivo destino (la plantilla)
                    )
                    logger.info(f"Hoja '{hoja_seleccionada}' copiada a 'Reemplazable' en plantilla '{nombre_plantilla}'.")
            except Exception as e:
                logger.error(f"Error copiando hoja a 'Reemplazable' para plantilla '{nombre_plantilla}': {e}", exc_info=True)


            # --- Actualizar Plantilla (Fórmulas, etc.) ---
            try:
                logger.info(f"Actualizando plantilla '{nombre_plantilla}' (fórmulas, etc.)...")
                # El código original obtenía el spreadsheet aquí, pero no parecía usarlo directamente
                # spreadsheet = patoba.obtener_g_sheet_por_id(id_archivo_plantilla)
                # Llamar a la función que actualiza la plantilla
                patoba.actualizar_plantilla(id_archivo_plantilla, planilla_lista) # Pasar ID y objeto planilla
                logger.info(f"Plantilla '{nombre_plantilla}' actualizada.")
            except Exception as e:
                 logger.error(f"Error actualizando plantilla '{nombre_plantilla}': {e}", exc_info=True)


            # --- Limpieza y Marcado Final ---
            try:
                # Marcar planilla como no lista en BD
                planilla_lista.listo = False
                planilla_lista.save()
                logger.info(f"Planilla {planilla_lista.descripcion} marcada como 'listo=False'.")

                # Guardar mapeo para referencia (si es necesario)
                # logger.info(f"Nombre descargable: {nombre_descargable}, Identificador plantilla: {id_archivo_plantilla}") # Reemplazo print
                mi_diccionario[str(nombre_descargable)] = str(id_archivo_plantilla)

                # Borrar el archivo original descargado de Drive
                logger.info(f"Borrando archivo original descargado de Drive (ID: {planilla_lista.identificador})...")
                patoba.borrar_por_id(planilla_lista.identificador)
                logger.info(f"Archivo original (ID: {planilla_lista.identificador}) borrado.")

                # Marcar proveedor asociado para procesamiento CSV (si no se hizo antes)
                # Esto parece redundante si buscar_modificar_registros ya procesó el CSV
                # Pero lo mantenemos si la lógica original lo tenía.
                if proveedor_obj:
                     lista_prov_obj = ListaProveedores.objects.get(id=proveedor_obj.identificador_id) # Obtener ListaProveedor asociada
                     if not lista_prov_obj.hay_csv_pendiente: # Marcar solo si no estaba ya pendiente
                         lista_prov_obj.hay_csv_pendiente = True
                         lista_prov_obj.save()
                         logger.info(f"Proveedor '{nombre_plantilla}' marcado con 'hay_csv_pendiente=True'.")
                # logger.info(f"Finalizando procesamiento para: {nombre_plantilla}") # Reemplazo print

            except Exception as e:
                logger.error(f"Error durante la limpieza final para {planilla_lista.descripcion}: {e}", exc_info=True)

        logger.info("Bucle principal de procesamiento de planillas finalizado.")
        # logger.debug(f"Diccionario final de mapeos: {mi_diccionario}") # Si es útil

    except Exception as e:
        logger.error(f"Error general en el bucle principal de procesamiento de planillas listas: {e}", exc_info=True)

    # Reemplazo de print("fin1")
    logger.info("Proceso principal finalizado.")


# --- Bloque de Ejecución Principal ---
if __name__ == '__main__':
    # Configura Django si es necesario (ej. si se ejecuta fuera de manage.py)
    # import django
    # os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'tu_proyecto.settings')
    # django.setup()
    logger.info("Script ejecutado directamente.")
    principal()


