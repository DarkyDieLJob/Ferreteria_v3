import os
os.environ['DJANGO_SETTINGS_MODULE'] = 'core_config.settings'
import django
django.setup()





import io
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
import pandas as pd

from openpyxl.writer.excel import save_virtual_workbook
from bdd.models import Listado_Planillas, Item, Sub_Carpeta, Sub_Titulo, ListaProveedores
from bdd.classes import Patoba
from bdd.funtions import get_emails
from django.conf import settings as cosnt

from asgiref.sync import sync_to_async

from django.conf import settings
import json
import os
import openpyxl
import csv

def registrar_log(texto):
    with open('log.txt', 'a') as f:
        print(texto, file=f)

def crear_o_actualizar_registro(row):
    try:
        # Recuperar o crear una instancia de Sub_Carpeta
        sub_carpeta_nombre = row.pop('sub_carpeta')
        sub_carpeta, created = Sub_Carpeta.objects.get_or_create(nombre=sub_carpeta_nombre)
        row['sub_carpeta'] = sub_carpeta
        sub_carpeta = Sub_Carpeta.objects.get(nombre=sub_carpeta_nombre)
        # Recuperar o crear una instancia de Sub_Titulo
        sub_titulo_nombre = row.pop('sub_titulo')
        #sub_titulo = Sub_Titulo.objects.get(nombre=sub_titulo_nombre)
        sub_titulo, created = Sub_Titulo.objects.get_or_create(nombre=sub_titulo_nombre)
        row['sub_titulo'] = sub_titulo
    except Exception as e:
        print(e)
    #print(row)
    try:
        item, created = Item.objects.get_or_create(codigo=row['codigo'], defaults=row)
        #item.calcular_precio_final()
        #item.calcular_precio_efectivo_final()
        #item.calcular_precio_rollo_final()
        #item.calcular_precio_rollo_efectivo_final()
        if not created:
            for key, value in row.items():
                #print(item, key, value)
                setattr(item, key, value)
                item.save()
    except Exception as e:
        print(e)
    try:
        item.marcar_actualizado()
        item.save()
    except Exception as e:
        print("item no se guardo porque el boludo tenia un error.",row)
        print(e)
    #print('creado o guardado: ', item)

def desactualizar_anteriores(filtro):
    anteriores = Item.objects.filter(codigo__endswith=filtro)
    for item in anteriores:
        item.marcar_desactualizado()
        item.save()

def buscar_modificar_registros(csv_file, filtro):
    # Cargar el archivo CSV
    desactualizar_anteriores(filtro)
    print('desactualizado: ',filtro)
    with open(csv_file, newline='') as csvfile:

        reader = csv.DictReader(csvfile)
        for row in reader:
            print(row)
            # Filtrar filas con valores nulos o vacíos
            if not (row['precio_base'] is None or row['precio_base'] == '' or row['precio_base'] == ' ' or row['precio_base'] == '-' or row['precio_base'] == '.' or row['precio_base'] == '#N/A' or row['precio_base'] == '#VALUE!'):
                crear_o_actualizar_registro(row)
    print('cargado: ', filtro)

def f():
    try:
        planilla = Listado_Planillas.objects.filter(proveedor="Caramia")
        # Obtener el archivo de Google Drive
        file_id = planilla.identificador
        request = drive_service.files().get_media(fileId=file_id)
        file = io.BytesIO()
        downloader = MediaIoBaseDownload(file, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()

        # Cargar el archivo xlsx en un objeto Workbook de openpyxl
        file.seek(0)
        wb = openpyxl.load_workbook(file)

        # Fusionar todas las hojas en una sola
        df = pd.concat([pd.DataFrame(ws.values) for ws in wb.worksheets])

        # Crear una nueva hoja llamada "cargar datos"
        ws = wb.create_sheet(title="cargar datos")

        # Guardar el DataFrame fusionado en la hoja "cargar datos"
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

        # Guardar el archivo xlsx modificado en Google Drive
        file.seek(0)
        media = MediaIoBaseUpload(file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        drive_service.files().update(fileId=file_id, media_body=media).execute()

        # Actualizar el valor de planilla.hoja a "cargar datos"
        planilla.hoja = "cargar datos"
        planilla.save()
    except Exception as e:
        print("no se encontro caramia {}".format(e))

def fusionar_hojas(file):
    # Cargar el archivo xlsx en un objeto Workbook de openpyxl
    wb = openpyxl.load_workbook(file)

    # Fusionar todas las hojas en una sola
    df = pd.concat([pd.DataFrame(ws.values) for ws in wb.worksheets if ws.title != "general"])

    # Crear una nueva hoja llamada "cargar datos"
    ws = wb.create_sheet(title="cargar datos")

    # Convertir el DataFrame en un arreglo de NumPy
    data = df.to_numpy()

    # Agregar los datos a la hoja de cálculo
    for row in data:
        ws.append(row.tolist())

    # Retornar el objeto Workbook con el contenido de la hoja "cargar datos"
    return wb

from x_cartel.models import Carteles, CartelesCajon

def marcar_revisar_carteles(id_proveedor):
    Carteles.objects.filter(proveedor=id_proveedor).update(revisar=True)
    CartelesCajon.objects.filter(proveedor=id_proveedor).update(revisar=True)
    print("Marcando carteles:", id_proveedor)

def principal():

    patoba = Patoba(None)

    gmail_service = patoba.gmail_service
    drive_service = patoba.drive_service

    get_emails(gmail_service, drive_service)
    #---------------------------------------------------------------------------
    items = patoba.listar(100, cosnt.INBOX)

    for i in items:
        try:
            Listado_Planillas.objects.get_or_create(
                descripcion=i['name'], identificador=i['id'])
        except:
            print('No se pudo get_or_create [{}] , [{}]'.format(i['name'],i['id']), flush=True)


    try:
        # Obtener el archivo de Google Drive
        planilla = Listado_Planillas.objects.filter(proveedor__text_display="Caramia")[0]
        file_id = planilla.identificador
        request = drive_service.files().get_media(fileId=file_id)
        file = io.BytesIO()
        downloader = MediaIoBaseDownload(file, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()

        # Cargar el archivo xlsx en un objeto Workbook de openpyxl
        file.seek(0)
        wb = fusionar_hojas(file)

        # Guardar los cambios en el archivo de Excel en Google Drive
        media = MediaIoBaseUpload(io.BytesIO(save_virtual_workbook(wb)), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        drive_service.files().update(fileId=file_id, media_body=media).execute()
    except Exception as e:
        print("no se encontro caramia {}".format(e), flush=True)



    lista = Listado_Planillas.objects.filter(listo=True)
    datos = Listado_Planillas.objects.filter(listo=False, descargar=False)

    for dato in datos:
            sheet_names = ["",]
            # Descargar el archivo de Excel desde Google Drive
            request = drive_service.files().get_media(fileId=dato.identificador)
            file = io.BytesIO()
            downloader = MediaIoBaseDownload(file, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()

            # Leer el contenido del archivo de Excel
            file.seek(0)
            try:
                #pd = "x"
                xls = pd.read_excel(file, sheet_name=None)

                # Obtener los nombres de las hojas
                sheet_names.extend(xls.keys())
                dato.hojas = ';'.join(sheet_names)
                print(dato.hojas, flush=True)
                dato.save()
            except Exception as e:
                print("error con pandas {}".format(e), flush=True)
    #---------------------------------------------------------------------------

    mi_diccionario = {}

    for sp in lista:
        nombre_proveedor = sp.descripcion
        nombre_plantilla = sp.proveedor
        print(nombre_plantilla)
        if nombre_plantilla == None or nombre_plantilla == 'Otros':
            continue
        nombre_descargable = "{}-{}".format(sp.proveedor, sp.fecha)
        hoja_seleccionada = sp.hoja
        print(hoja_seleccionada)

        id_archivo_proveedor = patoba.obtener_id_por_nombre(
            nombre_proveedor, cosnt.INBOX)

        id_archivo_plantilla = patoba.obtener_id_por_nombre(
            nombre_plantilla, cosnt.PLANTILLAS)

        sp.id_sp = id_archivo_plantilla
        sp.save()

        id_hoja_reemplazable = patoba.obtener_id_hoja_por_nombre(
            "Reemplazable", id_archivo_plantilla)
        try:
            # Obtener los datos de la hoja de cálculo
            result = patoba.sheet_service.spreadsheets().values().get(
                spreadsheetId=id_archivo_plantilla, range='BDD').execute()
            values = result.get('values', [])

            print(sp.proveedor.identificador.abreviatura)

            # Guardar los datos en un archivo CSV
            with open(f'{nombre_plantilla}.csv', 'w') as f:
                print('aqui 1', flush=True)
                writer = csv.writer(f)
                print('aqui 2', flush=True)
                writer.writerows(values)
            print('aqui 3', flush=True)
            # Cargar el archivo CSV

            #buscar_modificar_registros(f'{nombre_plantilla}.csv', sp.proveedor.identificador.abreviatura)
            csv_file =f'{nombre_plantilla}.csv'
            filtro = sp.proveedor.identificador.abreviatura
            print(csv_file,filtro)
            desactualizar_anteriores(filtro)
            marcar_revisar_carteles(sp.proveedor.id)
            print('desactualizado: ',filtro, flush=True)
            '''
            with open(csv_file, newline='') as csvfile:

                reader = csv.DictReader(csvfile)
                for row in reader:
                    print(row, flush=True)
                    # Filtrar filas con valores nulos o vacíos
                    if not (row['precio_base'] is None or row['precio_base'] == '' or row['precio_base'] == ' ' or row['precio_base'] == '-' or row['precio_base'] == '.' or row['precio_base'] == '#N/A' or row['precio_base'] == '#VALUE!'):
                        crear_o_actualizar_registro(row)
            print('cargado: ', filtro, flush=True)
            '''
        except Exception as e:
            print("No se encontro BDD", flush=True)
            print(e)

        # Copiar reemplazable
        if id_archivo_proveedor:
            request = patoba.drive_service.files().get_media(fileId=id_archivo_proveedor)
            hoja_proveedor = io.BytesIO()
            downloader = MediaIoBaseDownload(hoja_proveedor, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()
            hoja_proveedor.seek(0)
            hoja_proveedor = pd.read_excel(
                hoja_proveedor, sheet_name=None)
        else:
            print(
                f"No se encontró el archivo {nombre_proveedor}.xlsx en la carpeta Inbox", flush=True)
            hoja_proveedor = None

        print(type(hoja_proveedor))

        patoba.copiar_reemplazable(
            id_archivo_proveedor, hoja_seleccionada, id_hoja_reemplazable, id_archivo_plantilla)

        # Crear copia en descarga
        identificador = patoba.obtener_id_por_nombre(
            nombre_plantilla, cosnt.PLANTILLAS)

        print("nombre: {}, identificador: {}".format(
            nombre_descargable, identificador))
        mi_diccionario[str(nombre_descargable)] = str(identificador)

        spreadsheet = patoba.obtener_g_sheet_por_id(identificador)

        patoba.actualizar_plantilla(spreadsheet, sp)

        sp.listo = False
        sp.save()

        patoba.borrar_por_id(sp.identificador)
        print(nombre_plantilla)
        lista_proveedor = ListaProveedores.objects.get(nombre=nombre_plantilla)
        lista_proveedor.hay_csv_pendiente = True
        lista_proveedor.save()

    print("fin")

if __name__ == '__main__':
    principal()

