
from .outsider import arrancar_django_config
arrancar_django_config()

import openpyxl


import openpyxl
from datetime import datetime

def queryset_to_xlsx(queryset, filename):
    ''' Escribir un queryset en un archivo xlsx '''
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # Verificar si el queryset contiene diccionarios o objetos de modelo
    if isinstance(queryset.first(), dict):
        # Escribir encabezados para diccionarios
        headers = queryset.first().keys()
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = len(header)
            sheet.cell(row=1, column=col_num).alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # Escribir datos para diccionarios
        for row_num, row in enumerate(queryset, 2):
            for col_num, (key, value) in enumerate(row.items(), 1):
                if isinstance(value, datetime):
                    value = value.replace(tzinfo=None)  # Eliminar la información de zona horaria
                try:
                    sheet.cell(row=row_num, column=col_num).value = value
                except ValueError:
                    sheet.cell(row=row_num, column=col_num).value = str(value)
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = max(sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width, len(str(value)))
                sheet.cell(row=row_num, column=col_num).alignment = openpyxl.styles.Alignment(horizontal='center')
    else:
        # Escribir encabezados para objetos de modelo
        headers = queryset.model._meta.fields
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header.name
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = len(header.name)
            sheet.cell(row=1, column=col_num).alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # Escribir datos para objetos de modelo
        for row_num, row in enumerate(queryset, 2):
            for col_num, field in enumerate(headers, 1):
                value = getattr(row, field.name)  # Use getattr() to access the field value
                if isinstance(value, datetime):
                    value = value.replace(tzinfo=None)  # Eliminar la información de zona horaria
                try:
                    sheet.cell(row=row_num, column=col_num).value = value
                except ValueError:
                    sheet.cell(row=row_num, column=col_num).value = str(value)
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = max(sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width, len(str(value)))
                sheet.cell(row=row_num, column=col_num).alignment = openpyxl.styles.Alignment(horizontal='center')
    
    workbook.save(filename)
    
def queryset_to_sheet(queryset, sheet):
    ''' Escribir un queryset en una hoja de un archivo de sheets '''
    headers = queryset.model._meta.fields
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header.name
    
    for row_num, row in enumerate(queryset, 2):
        for col_num, field in enumerate(headers, 1):
            value = getattr(row, field.name)
            sheet.cell(row=row_num, column=col_num).value = value
            
def diccionario_to_xlsx(diccionario, filename):
    ''' Escribir un diccionario en un archivo xlsx '''
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # Write headers
    headers = diccionario.keys()
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = len(header)
        sheet.cell(row=1, column=col_num).alignment = openpyxl.styles.Alignment(horizontal='center')
    
    # Write data
    for row_num, key in enumerate(diccionario, 2):
        sheet.cell(row=row_num, column=1).value = key
        sheet.cell(row=row_num, column=2).value = diccionario[key]
        sheet.column_dimensions['A'].width = max(sheet.column_dimensions['A'].width, len(str(key)))
        sheet.column_dimensions['B'].width = max(sheet.column_dimensions['B'].width, len(str(diccionario[key])))
        sheet.cell(row=row_num, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row=row_num, column=2).alignment = openpyxl.styles.Alignment(horizontal='center')
    
    workbook.save(filename)
            
from bdd.models import Lista_Pedidos, Proveedor
from facturacion.models import ArticuloVendido
from django.conf import settings
from django.db.models import Count, Sum
from django.db.models.functions import TruncMonth
def ejecutar():
    queryset = Lista_Pedidos.objects.all()
    queryset_to_xlsx(queryset, f'{settings.MEDIA_ROOT}/pedidos/pedidos.xlsx')
    
    # Contar los articulos vendidos
    articulos_vendidos_contados = ArticuloVendido.objects.annotate(
        month=TruncMonth('transaccion__fecha')
    ).values(
        'month', 'item__codigo', 'item__descripcion'
    ).annotate(
        count=Count('item'),
        total_quantity=Sum('cantidad')
    ).order_by('month', 'item__codigo')
    
    filename = f'{settings.MEDIA_ROOT}/pedidos/articulos_vendidos.xlsx'
    queryset_to_xlsx(articulos_vendidos_contados, filename)

    proveedores = queryset.values_list('proveedor', flat=True).distinct()

    for proveedor_id in proveedores:
        filtro = {'proveedor': proveedor_id}
        queryset_filtrado = queryset.filter(**filtro)
        proveedor = Proveedor.objects.get(pk=proveedor_id)
        filename = f'{settings.MEDIA_ROOT}/pedidos/{proveedor.text_display}.xlsx'
        queryset_to_xlsx(queryset_filtrado, filename)
    
    

if __name__ == '__main__':
    ejecutar()